#!/usr/bin/env python3
"""
Generate Complete Supplier RFQ Package
======================================

This script generates BOTH:
1. Professional PDF RFQ (using generate_rfq_pdf.py)
2. Professional Excel file with items and instructions

Usage:
    python3 generate_supplier_rfq.py <config.json>

The config JSON should include:
- All standard RFQ fields (company, rfq_details, etc.)
- items_data: array of items OR items_file: path to JSON file with items
"""

import json
import sys
import os
import subprocess
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_excel_from_config(config, base_path):
    """Generate professional Excel file from RFQ config"""
    
    # Get items data
    items = None
    if 'items_data' in config:
        items = config['items_data']
    elif 'items_file' in config:
        items_file = os.path.join(base_path, config['items_file'])
        with open(items_file, 'r') as f:
            items = json.load(f)
    elif 'items' in config:
        # Items might be in simplified format in config
        items = config['items']
    
    if not items:
        print("⚠ No items data found - skipping Excel generation")
        return None
    
    # Extract config values
    company = config['company']
    rfq = config['rfq_details']
    colors = config.get('colors', {
        'primary': '#D97706',
        'accent': '#0F172A',
        'text': '#374151'
    })
    
    # Get Excel-specific config or use defaults
    excel_config = config.get('excel', {})
    
    # Remove # from color codes
    primary_color = colors['primary'].replace('#', '')
    secondary_color = colors['accent'].replace('#', '')
    instruction_color = excel_config.get('instruction_color', 'FEF3C7')
    light_gray = excel_config.get('light_gray', 'F3F4F6')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RFQ"
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 30
    
    # Define styles
    header_font = Font(name='Avenir', size=16, bold=True, color='FFFFFF')
    subheader_font = Font(name='Avenir', size=11, bold=True, color='FFFFFF')
    contact_font = Font(name='Avenir', size=9, color='000000')
    column_header_font = Font(name='Avenir', size=10, bold=True, color='FFFFFF')
    data_font = Font(name='Avenir', size=9)
    
    # Color fills
    primary_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type='solid')
    secondary_fill = PatternFill(start_color=secondary_color, end_color=secondary_color, fill_type='solid')
    instruction_fill = PatternFill(start_color=instruction_color, end_color=instruction_color, fill_type='solid')
    light_gray_fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type='solid')
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get certification text
    cert_text = excel_config.get('certification', 'EDWOSB CERTIFIED')
    
    # Row 1: Company Header
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = f"{company['name']} - {cert_text}"
    cell.font = header_font
    cell.fill = primary_fill
    cell.alignment = center_align
    ws.row_dimensions[1].height = 25
    
    # Row 2: RFQ Title
    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = f"REQUEST FOR QUOTE - {rfq['title']} ({len(items)} Items)"
    cell.font = subheader_font
    cell.fill = secondary_fill
    cell.alignment = center_align
    ws.row_dimensions[2].height = 20
    
    # Row 3: Contact Info
    ws.merge_cells('A3:J3')
    cell = ws['A3']
    cell.value = f"{company['address']} | Phone: {company['phone']} | Email: {company['email']}"
    cell.font = contact_font
    cell.fill = light_gray_fill
    cell.alignment = center_align
    ws.row_dimensions[3].height = 18
    
    # Row 4: RFQ Details
    ws.merge_cells('A4:J4')
    cell = ws['A4']
    cell.value = f"RFQ #: {rfq['rfq_number']} | Due Date: {rfq['due_date']} {rfq.get('due_time', '')} | Project: {rfq.get('project_name', 'Client Project')}"
    cell.font = Font(name='Avenir', size=10, bold=True)
    cell.fill = light_gray_fill
    cell.alignment = center_align
    ws.row_dimensions[4].height = 18
    
    # Row 5: Blank
    ws.row_dimensions[5].height = 8
    
    # Rows 6+: INSTRUCTIONS
    instructions = excel_config.get('instructions', [
        '⚠️ INSTRUCTIONS FOR SUPPLIERS - PLEASE READ CAREFULLY',
        f'1. FILL IN COLUMNS F through J for each of the {len(items)} items below',
        '2. Column F: Your Part Number (or equivalent product)',
        '3. Column G: Your Unit Price (government/volume pricing - NOT retail)',
        '4. Column H: Extended Total (Unit Price × Quantity)',
        '5. Column I: Lead Time (in business days or weeks)',
        '6. Column J: Notes (substitutions, approved equivalents, volume discounts)',
        '7. ADD FREIGHT COST as a separate line item at the bottom',
        '8. CONFIRM TAX EXEMPTION applied (government client - sales tax exempt)',
        f'9. EMAIL completed file to: {company["email"]} by {rfq["due_date"]} {rfq.get("due_time", "")}'
    ])
    
    current_row = 6
    for idx, instruction in enumerate(instructions):
        ws.merge_cells(f'A{current_row}:J{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = instruction
        if idx == 0:
            cell.font = Font(name='Avenir', size=11, bold=True, color='000000')
        else:
            cell.font = Font(name='Avenir', size=9, bold=False, color='000000')
        cell.fill = instruction_fill
        cell.alignment = left_align
        ws.row_dimensions[current_row].height = 18
        current_row += 1
    
    # Blank row
    ws.row_dimensions[current_row].height = 8
    current_row += 1
    
    # Column Headers
    headers = excel_config.get('column_headers', [
        'Item #', 'Description', 'Specification', 'Quantity', 'Unit',
        'YOUR PART NUMBER', 'YOUR UNIT PRICE', 'EXTENDED TOTAL', 'LEAD TIME', 'NOTES'
    ])
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = column_header_font
        cell.fill = secondary_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    ws.row_dimensions[current_row].height = 30
    header_row = current_row
    current_row += 1
    
    # Add items
    for idx, item in enumerate(items, start=1):
        # Item number
        cell = ws.cell(row=current_row, column=1)
        cell.value = idx
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border
        
        # Description
        cell = ws.cell(row=current_row, column=2)
        # Handle different item formats
        desc = item.get('description', item.get('item_description', ''))
        cell.value = desc
        cell.font = data_font
        cell.alignment = left_align
        cell.border = thin_border
        
        # Specification
        cell = ws.cell(row=current_row, column=3)
        spec = item.get('spec', item.get('specifications', item.get('specification', '')))
        cell.value = spec
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border
        
        # Quantity
        cell = ws.cell(row=current_row, column=4)
        qty = item.get('quantity', item.get('estimated_quantity', ''))
        cell.value = qty
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border
        
        # Unit
        cell = ws.cell(row=current_row, column=5)
        unit = item.get('unit', 'each')
        cell.value = unit
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border
        
        # Blank columns for supplier (F-J)
        for col in range(6, 11):
            cell = ws.cell(row=current_row, column=col)
            cell.font = data_font
            cell.border = thin_border
            if col == 7 or col == 8:
                cell.alignment = right_align
            else:
                cell.alignment = left_align
        
        ws.row_dimensions[current_row].height = 30
        current_row += 1
    
    # Footer rows
    current_row += 1
    delivery_location = excel_config.get('delivery_location', 'Client Location')
    
    # Freight row
    ws.merge_cells(f'A{current_row}:E{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = f'FREIGHT COST to {delivery_location}'
    cell.font = Font(name='Avenir', size=10, bold=True)
    cell.fill = instruction_fill
    cell.alignment = left_align
    cell.border = thin_border
    
    cell = ws.cell(row=current_row, column=8)
    cell.value = '(Enter freight cost here)'
    cell.font = Font(name='Avenir', size=9, italic=True)
    cell.fill = instruction_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.row_dimensions[current_row].height = 25
    
    # Total row
    current_row += 1
    ws.merge_cells(f'A{current_row}:E{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = 'TOTAL PROJECT COST (Items + Freight)'
    cell.font = Font(name='Avenir', size=11, bold=True)
    cell.fill = primary_fill
    cell.alignment = left_align
    cell.border = thin_border
    
    cell = ws.cell(row=current_row, column=8)
    cell.value = '(Enter total here)'
    cell.font = Font(name='Avenir', size=10, bold=True)
    cell.fill = primary_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.row_dimensions[current_row].height = 25
    
    # Freeze panes
    ws.freeze_panes = f'A{header_row + 1}'
    
    # Save Excel file
    excel_filename = excel_config.get('output_file')
    if not excel_filename:
        # Generate default filename
        excel_filename = rfq['rfq_number'].replace('-', '_') + '_ITEMS.xlsx'
    
    excel_path = os.path.join(base_path, excel_filename)
    wb.save(excel_path)
    
    return excel_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 generate_supplier_rfq.py <config.json>")
        print("\nGenerates:")
        print("  1. Professional PDF RFQ")
        print("  2. Professional Excel file with items")
        sys.exit(1)
    
    config_file = sys.argv[1]
    base_path = os.path.dirname(config_file) or '.'
    
    # Load config
    with open(config_file, 'r') as f:
        config = json.load(f)
    
    print(f"📋 Generating Supplier RFQ Package from {config_file}...")
    print()
    
    # Generate PDF (using existing script)
    print("1️⃣ Generating PDF RFQ...")
    pdf_result = subprocess.run(
        ['python3', 'generate_rfq_pdf.py', config_file],
        capture_output=True,
        text=True
    )
    if pdf_result.returncode == 0:
        print(f"   ✓ {pdf_result.stdout.strip()}")
    else:
        print(f"   ⚠ PDF generation had issues: {pdf_result.stderr}")
    
    # Generate Excel
    print("2️⃣ Generating Excel file...")
    excel_path = generate_excel_from_config(config, base_path)
    if excel_path:
        print(f"   ✓ Generated: {excel_path}")
    else:
        print(f"   ⚠ Excel generation skipped (no items data)")
    
    print()
    print("✅ RFQ Package Complete!")
    print(f"📁 Location: {base_path}")
    print()
    print("📤 Ready to send to suppliers:")
    print("   - PDF RFQ (professional quote request)")
    print("   - Excel file (for supplier to fill in pricing)")
    print("   - Specifications PDF (if applicable)")


if __name__ == "__main__":
    main()
