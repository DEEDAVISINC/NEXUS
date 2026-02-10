#!/usr/bin/env python3
"""
TEMPLATE: Create Professional Supplier RFQ Excel File
=====================================================

This template creates a supplier-safe Excel file with:
- Company branding and contact info
- Clear step-by-step instructions for suppliers
- Professional formatting with customizable colors
- All items listed with blank columns for supplier pricing

USAGE:
1. Copy this file to your project folder
2. Update the CUSTOMIZATION section below
3. Prepare your items data as JSON (see format below)
4. Run: python3 create_supplier_rfq_excel.py

REQUIRED JSON FORMAT (save as items.json or similar):
[
  {
    "item": 1,
    "description": "Item description here",
    "spec": "Specification here",
    "quantity": 100
  },
  ...
]
"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CUSTOMIZATION SECTION - EDIT THESE VALUES FOR EACH RFQ
# ============================================================================

# File paths
INPUT_JSON_FILE = 'clean_items_for_suppliers.json'  # Your items data file
OUTPUT_EXCEL_FILE = 'DEE_DAVIS_INC_TRAFFIC_SIGNS_RFQ_109_ITEMS.xlsx'  # Output filename

# Company Information
COMPANY_NAME = "DEE DAVIS INC"
COMPANY_CERTIFICATION = "EDWOSB CERTIFIED"  # Or "WOSB CERTIFIED", "SBA CERTIFIED", etc.
COMPANY_ADDRESS = "755 W Big Beaver Rd, Suite 2020, Troy, MI 48084"
COMPANY_PHONE = "248-376-4550"
COMPANY_EMAIL = "info@deedavis.biz"

# RFQ Details
RFQ_NUMBER = "DDI-2026-007"
RFQ_TITLE = "TRAFFIC SIGNS"
TOTAL_ITEMS = 109
DUE_DATE = "February 10, 2026 @ 12:00 PM EST"
PROJECT_NAME = "Michigan Municipal Road Commission"

# Column Headers (customize as needed)
COLUMN_HEADERS = [
    'Item #',
    'Description', 
    'ASTM Specification',  # Change to "Specification", "Model", etc. as needed
    'Quantity',
    'Unit',
    'YOUR PART NUMBER',
    'YOUR UNIT PRICE',
    'EXTENDED TOTAL',
    'LEAD TIME',
    'NOTES'
]

# Colors (change these to customize the look)
# Use hex color codes without the # symbol
PRIMARY_COLOR = 'D97706'      # Orange - for main headers
SECONDARY_COLOR = '0F172A'    # Dark Blue - for subheaders and column headers
INSTRUCTION_COLOR = 'FEF3C7'  # Light Yellow - for instruction section
LIGHT_GRAY = 'F3F4F6'         # Light Gray - for info rows

# Instructions for suppliers (customize as needed)
INSTRUCTIONS = [
    '⚠️ INSTRUCTIONS FOR SUPPLIERS - PLEASE READ CAREFULLY',
    '1. FILL IN COLUMNS F through J for each of the {} items below'.format(TOTAL_ITEMS),
    '2. Column F: Your Part Number (or equivalent MUTCD-compliant product)',
    '3. Column G: Your Unit Price (government/volume pricing - NOT retail)',
    '4. Column H: Extended Total (Unit Price × Quantity)',
    '5. Column I: Lead Time (in business days or weeks)',
    '6. Column J: Notes (substitutions, approved equivalents, volume discounts applied)',
    '7. ADD FREIGHT COST as a separate line item at the bottom of the sheet',
    '8. CONFIRM TAX EXEMPTION applied (Michigan government client - sales tax exempt)',
    '9. EMAIL completed file + specs to: {} by {}'.format(COMPANY_EMAIL, DUE_DATE)
]

# Delivery/Shipping Location (for freight calculation)
DELIVERY_LOCATION = "Waterford, MI 48328"

# ============================================================================
# END CUSTOMIZATION SECTION
# ============================================================================

# Load items data
with open(INPUT_JSON_FILE, 'r') as f:
    items = json.load(f)

# Create new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "RFQ"

# Set column widths
ws.column_dimensions['A'].width = 8   # Item #
ws.column_dimensions['B'].width = 50  # Description
ws.column_dimensions['C'].width = 20  # Specification
ws.column_dimensions['D'].width = 10  # Quantity
ws.column_dimensions['E'].width = 8   # Unit
ws.column_dimensions['F'].width = 18  # Your Part Number
ws.column_dimensions['G'].width = 12  # Unit Price
ws.column_dimensions['H'].width = 14  # Extended Total
ws.column_dimensions['I'].width = 12  # Lead Time
ws.column_dimensions['J'].width = 30  # Notes

# Define styles
header_font = Font(name='Avenir', size=16, bold=True, color='FFFFFF')
subheader_font = Font(name='Avenir', size=11, bold=True, color='FFFFFF')
contact_font = Font(name='Avenir', size=9, color='000000')
column_header_font = Font(name='Avenir', size=10, bold=True, color='FFFFFF')
data_font = Font(name='Avenir', size=9)

# Color fills
primary_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type='solid')
secondary_fill = PatternFill(start_color=SECONDARY_COLOR, end_color=SECONDARY_COLOR, fill_type='solid')
instruction_fill = PatternFill(start_color=INSTRUCTION_COLOR, end_color=INSTRUCTION_COLOR, fill_type='solid')
light_gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Row 1: Company Header
ws.merge_cells('A1:J1')
cell = ws['A1']
cell.value = f'{COMPANY_NAME} - {COMPANY_CERTIFICATION}'
cell.font = header_font
cell.fill = primary_fill
cell.alignment = center_align
ws.row_dimensions[1].height = 25

# Row 2: RFQ Title
ws.merge_cells('A2:J2')
cell = ws['A2']
cell.value = f'REQUEST FOR QUOTE - {RFQ_TITLE} ({TOTAL_ITEMS} Items)'
cell.font = subheader_font
cell.fill = secondary_fill
cell.alignment = center_align
ws.row_dimensions[2].height = 20

# Row 3: Contact Info
ws.merge_cells('A3:J3')
cell = ws['A3']
cell.value = f'{COMPANY_ADDRESS} | Phone: {COMPANY_PHONE} | Email: {COMPANY_EMAIL}'
cell.font = contact_font
cell.fill = light_gray_fill
cell.alignment = center_align
ws.row_dimensions[3].height = 18

# Row 4: RFQ Details
ws.merge_cells('A4:J4')
cell = ws['A4']
cell.value = f'RFQ #: {RFQ_NUMBER} | Due Date: {DUE_DATE} | Project: {PROJECT_NAME}'
cell.font = Font(name='Avenir', size=10, bold=True)
cell.fill = light_gray_fill
cell.alignment = center_align
ws.row_dimensions[4].height = 18

# Row 5: Blank
ws.row_dimensions[5].height = 8

# Rows 6+: INSTRUCTIONS (highlighted section)
current_row = 6
for idx, instruction in enumerate(INSTRUCTIONS):
    ws.merge_cells(f'A{current_row}:J{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = instruction
    if idx == 0:
        cell.font = Font(name='Avenir', size=11, bold=True, color='000000')
    else:
        cell.font = Font(name='Avenir', size=9, bold=False, color='000000')
    cell.fill = instruction_fill
    cell.alignment = left_align
    ws.row_dimensions[current_row].height = 18
    current_row += 1

# Blank row after instructions
ws.row_dimensions[current_row].height = 8
current_row += 1

# Column Headers Row
for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
    cell = ws.cell(row=current_row, column=col_idx)
    cell.value = header
    cell.font = column_header_font
    cell.fill = secondary_fill
    cell.alignment = center_align
    cell.border = thin_border

ws.row_dimensions[current_row].height = 30
header_row = current_row
current_row += 1

# Add all items
for idx, item in enumerate(items, start=1):
    # Item number
    cell = ws.cell(row=current_row, column=1)
    cell.value = idx
    cell.font = data_font
    cell.alignment = center_align
    cell.border = thin_border
    
    # Description
    cell = ws.cell(row=current_row, column=2)
    cell.value = item['description']
    cell.font = data_font
    cell.alignment = left_align
    cell.border = thin_border
    
    # Specification
    cell = ws.cell(row=current_row, column=3)
    cell.value = item['spec']
    cell.font = data_font
    cell.alignment = center_align
    cell.border = thin_border
    
    # Quantity
    cell = ws.cell(row=current_row, column=4)
    cell.value = item['quantity']
    cell.font = data_font
    cell.alignment = center_align
    cell.border = thin_border
    
    # Unit
    cell = ws.cell(row=current_row, column=5)
    cell.value = 'each'
    cell.font = data_font
    cell.alignment = center_align
    cell.border = thin_border
    
    # Blank columns for supplier to fill (F, G, H, I, J)
    for col in range(6, 11):
        cell = ws.cell(row=current_row, column=col)
        cell.font = data_font
        cell.border = thin_border
        if col == 7 or col == 8:  # Price columns
            cell.alignment = right_align
        else:
            cell.alignment = left_align
    
    ws.row_dimensions[current_row].height = 30
    current_row += 1

# Add footer rows for totals and freight
current_row += 1
ws.merge_cells(f'A{current_row}:E{current_row}')
cell = ws[f'A{current_row}']
cell.value = f'FREIGHT COST to {DELIVERY_LOCATION}'
cell.font = Font(name='Avenir', size=10, bold=True)
cell.fill = instruction_fill
cell.alignment = left_align
cell.border = thin_border

cell = ws.cell(row=current_row, column=8)
cell.value = '(Enter freight cost here)'
cell.font = Font(name='Avenir', size=9, italic=True)
cell.fill = instruction_fill
cell.alignment = center_align
cell.border = thin_border
ws.row_dimensions[current_row].height = 25

current_row += 1
ws.merge_cells(f'A{current_row}:E{current_row}')
cell = ws[f'A{current_row}']
cell.value = 'TOTAL PROJECT COST (Items + Freight)'
cell.font = Font(name='Avenir', size=11, bold=True)
cell.fill = primary_fill
cell.alignment = left_align
cell.border = thin_border

cell = ws.cell(row=current_row, column=8)
cell.value = '(Enter total here)'
cell.font = Font(name='Avenir', size=10, bold=True)
cell.fill = primary_fill
cell.alignment = center_align
cell.border = thin_border
ws.row_dimensions[current_row].height = 25

# Freeze panes (freeze header rows)
ws.freeze_panes = f'A{header_row + 1}'

# Save file
wb.save(OUTPUT_EXCEL_FILE)
print(f"✓ Created professional RFQ Excel file: {OUTPUT_EXCEL_FILE}")
print(f"  - {len(items)} items with clear instructions")
print(f"  - Company branding and contact info")
print(f"  - Highlighted instruction section")
print(f"  - Supplier fill-in columns clearly marked")
print(f"  - Ready to send to suppliers!")
