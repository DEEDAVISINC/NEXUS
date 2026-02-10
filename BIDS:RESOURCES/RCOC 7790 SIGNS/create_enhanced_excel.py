#!/usr/bin/env python3
"""
Create enhanced supplier-safe Excel file with clear instructions and company identifiers
"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load the clean items data
with open('clean_items_for_suppliers.json', 'r') as f:
    items = json.load(f)

# Create new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Traffic Signs RFQ"

# Set column widths
ws.column_dimensions['A'].width = 8   # Item #
ws.column_dimensions['B'].width = 50  # Description
ws.column_dimensions['C'].width = 20  # ASTM Spec
ws.column_dimensions['D'].width = 10  # Quantity
ws.column_dimensions['E'].width = 8   # Unit
ws.column_dimensions['F'].width = 18  # Your Part Number
ws.column_dimensions['G'].width = 12  # Unit Price
ws.column_dimensions['H'].width = 14  # Extended Total
ws.column_dimensions['I'].width = 12  # Lead Time
ws.column_dimensions['J'].width = 30  # Notes

# Define styles
header_font = Font(name='Avenir', size=16, bold=True, color='FFFFFF')
subheader_font = Font(name='Avenir', size=11, bold=True, color='FFFFFF')
instruction_font = Font(name='Avenir', size=10, bold=True, color='000000')
contact_font = Font(name='Avenir', size=9, color='000000')
column_header_font = Font(name='Avenir', size=10, bold=True, color='FFFFFF')
data_font = Font(name='Avenir', size=9)

# Colors
orange_fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
dark_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
yellow_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
light_gray_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Row 1: Company Header
ws.merge_cells('A1:J1')
cell = ws['A1']
cell.value = 'DEE DAVIS INC - EDWOSB CERTIFIED'
cell.font = header_font
cell.fill = orange_fill
cell.alignment = center_align
ws.row_dimensions[1].height = 25

# Row 2: RFQ Title
ws.merge_cells('A2:J2')
cell = ws['A2']
cell.value = 'REQUEST FOR QUOTE - TRAFFIC SIGNS (109 Items)'
cell.font = subheader_font
cell.fill = dark_fill
cell.alignment = center_align
ws.row_dimensions[2].height = 20

# Row 3: Contact Info
ws.merge_cells('A3:J3')
cell = ws['A3']
cell.value = '755 W Big Beaver Rd, Suite 2020, Troy, MI 48084 | Phone: 248-376-4550 | Email: info@deedavis.biz'
cell.font = contact_font
cell.fill = light_gray_fill
cell.alignment = center_align
ws.row_dimensions[3].height = 18

# Row 4: RFQ Details
ws.merge_cells('A4:J4')
cell = ws['A4']
cell.value = 'RFQ #: DDI-2026-007 | Due Date: February 10, 2026 @ 12:00 PM EST | Project: Michigan Municipal Road Commission'
cell.font = Font(name='Avenir', size=10, bold=True)
cell.fill = light_gray_fill
cell.alignment = center_align
ws.row_dimensions[4].height = 18

# Row 5: Blank
ws.row_dimensions[5].height = 8

# Row 6-10: INSTRUCTIONS (Yellow highlighted section)
instructions = [
    '⚠️ INSTRUCTIONS FOR SUPPLIERS - PLEASE READ CAREFULLY',
    '1. FILL IN COLUMNS F through J for each of the 109 items below',
    '2. Column F: Your Part Number (or equivalent MUTCD-compliant product)',
    '3. Column G: Your Unit Price (government/volume pricing - NOT retail)',
    '4. Column H: Extended Total (Unit Price × Quantity)',
    '5. Column I: Lead Time (in business days or weeks)',
    '6. Column J: Notes (substitutions, approved equivalents, volume discounts applied)',
    '7. ADD FREIGHT COST as a separate line item at the bottom of the sheet',
    '8. CONFIRM TAX EXEMPTION applied (Michigan government client - sales tax exempt)',
    '9. EMAIL completed file + specs to: info@deedavis.biz by February 10, 2026 @ 12 PM EST'
]

current_row = 6
for instruction in instructions:
    ws.merge_cells(f'A{current_row}:J{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = instruction
    if current_row == 6:
        cell.font = Font(name='Avenir', size=11, bold=True, color='000000')
    else:
        cell.font = Font(name='Avenir', size=9, bold=False, color='000000')
    cell.fill = yellow_fill
    cell.alignment = left_align
    ws.row_dimensions[current_row].height = 18
    current_row += 1

# Row after instructions: Blank
ws.row_dimensions[current_row].height = 8
current_row += 1

# Column Headers Row
headers = ['Item #', 'Description', 'ASTM Specification', 'Quantity', 'Unit', 
           'YOUR PART NUMBER', 'YOUR UNIT PRICE', 'EXTENDED TOTAL', 'LEAD TIME', 'NOTES']

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=current_row, column=col_idx)
    cell.value = header
    cell.font = column_header_font
    cell.fill = dark_fill
    cell.alignment = center_align
    cell.border = thin_border

ws.row_dimensions[current_row].height = 30
header_row = current_row
current_row += 1

# Add all items
for idx, item in enumerate(items, start=1):
    # Item number
    cell = ws.cell(row=current_row, column=1)
    cell.value = idx
    cell.font = data_font
    cell.alignment = center_align
    cell.border = thin_border
    
    # Description
    cell = ws.cell(row=current_row, column=2)
    cell.value = item['description']
    cell.font = data_font
    cell.alignment = left_align
    cell.border = thin_border
    
    # ASTM Specification
    cell = ws.cell(row=current_row, column=3)
    cell.value = item['spec']
    cell.font = data_font
    cell.alignment = center_align
    cell.border = thin_border
    
    # Quantity
    cell = ws.cell(row=current_row, column=4)
    cell.value = item['quantity']
    cell.font = data_font
    cell.alignment = center_align
    cell.border = thin_border
    
    # Unit
    cell = ws.cell(row=current_row, column=5)
    cell.value = 'each'
    cell.font = data_font
    cell.alignment = center_align
    cell.border = thin_border
    
    # Blank columns for supplier to fill (F, G, H, I, J)
    for col in range(6, 11):
        cell = ws.cell(row=current_row, column=col)
        cell.font = data_font
        cell.border = thin_border
        if col == 7 or col == 8:  # Price columns
            cell.alignment = right_align
        else:
            cell.alignment = left_align
    
    ws.row_dimensions[current_row].height = 30
    current_row += 1

# Add footer rows for totals and freight
current_row += 1
ws.merge_cells(f'A{current_row}:E{current_row}')
cell = ws[f'A{current_row}']
cell.value = 'FREIGHT COST to Waterford, MI 48328'
cell.font = Font(name='Avenir', size=10, bold=True)
cell.fill = yellow_fill
cell.alignment = left_align
cell.border = thin_border

cell = ws.cell(row=current_row, column=8)
cell.value = '(Enter freight cost here)'
cell.font = Font(name='Avenir', size=9, italic=True)
cell.fill = yellow_fill
cell.alignment = center_align
cell.border = thin_border
ws.row_dimensions[current_row].height = 25

current_row += 1
ws.merge_cells(f'A{current_row}:E{current_row}')
cell = ws[f'A{current_row}']
cell.value = 'TOTAL PROJECT COST (Items + Freight)'
cell.font = Font(name='Avenir', size=11, bold=True)
cell.fill = orange_fill
cell.alignment = left_align
cell.border = thin_border

cell = ws.cell(row=current_row, column=8)
cell.value = '(Enter total here)'
cell.font = Font(name='Avenir', size=10, bold=True)
cell.fill = orange_fill
cell.alignment = center_align
cell.border = thin_border
ws.row_dimensions[current_row].height = 25

# Freeze panes (freeze header rows)
ws.freeze_panes = f'A{header_row + 1}'

# Save file
output_filename = 'DEE_DAVIS_INC_TRAFFIC_SIGNS_RFQ_109_ITEMS.xlsx'
wb.save(output_filename)
print(f"✓ Created enhanced Excel file: {output_filename}")
print(f"  - 109 items with clear instructions")
print(f"  - Company branding and contact info")
print(f"  - Yellow highlighted instruction section")
print(f"  - Supplier fill-in columns clearly marked")
