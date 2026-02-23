#!/usr/bin/env python3
"""
NEXUS CONTRACT INTELLIGENCE ENGINE
====================================
Ingests market intelligence data (expiring contracts, prime/sub directories),
cross-references them against DDI's service lanes, and surfaces three-avenue
opportunities:

  1. SUB UNDER PRIME — Get on existing prime's vendor list for immediate revenue
  2. PRIME THE RECOMPETE — When contract expires, DDI bids as prime with past performance
  3. HIRE SUBS — DDI primes the work, brings in Tier 2 subs from directory

Data Sources:
  - expiring-contracts XLSX (USASpending / GovCon Giants)
  - prime-contractor-directory PDF (SBLO contacts)
  - subcontractor-directory PDF (Tier 2 contacts)
  - Any future CSV/XLSX/PDF intelligence drops

This module is called by api_server.py and integrates with the agenda dashboard.
"""

import os
import re
import json
import csv
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import defaultdict
from dotenv import load_dotenv

load_dotenv()

BIDS_ROOT = Path(os.environ.get('BIDS_ROOT', '/Users/deedavis/NEXUS BACKEND/BIDS:RESOURCES'))
INTEL_DATA_PATH = Path(os.environ.get(
    'INTEL_DATA_PATH',
    '/Users/deedavis/NEXUS BACKEND/intelligence_data.json'
))

# DDI's actual service lanes — NAICS codes where DDI can prime or sub
DDI_SERVICE_LANES = {
    '561210': 'Facilities Support',
    '561720': 'Janitorial',
    '561730': 'Landscaping / Grounds',
    '561612': 'Security Guards',
    '561611': 'Investigation Services',
    '561110': 'Office Admin',
    '561320': 'Temp Staffing',
    '561990': 'Other Support Services',
    '561499': 'Business Support',
    '492110': 'Courier / Express Delivery',
    '492210': 'Local Messenger / Courier',
    '488510': 'Freight Transport Arrangement',
    '488999': 'Transport Support',
    '488190': 'Air Transport Support',
    '621511': 'Medical Labs / Drug Testing',
    '621999': 'Ambulance / Medical Services',
    '621498': 'Outpatient Care',
    '339112': 'Medical Instruments / Supplies',
    '541611': 'Admin Management Consulting',
    '541612': 'HR Consulting',
    '541614': 'Process / Logistics Consulting',
    '561421': 'Telephone Answering',
    '238990': 'Specialty Trade Contractors',
    '562111': 'Solid Waste Collection',
    '562998': 'Waste Management',
    '423840': 'Industrial Supplies (Wholesale)',
    '423850': 'Service Establishment Equip',
    '423990': 'Durable Goods (Wholesale)',
    '424120': 'Stationery / Office Supplies',
    '424690': 'Chemical Products (Wholesale)',
    '424910': 'Farm Supplies (Wholesale)',
    '444190': 'Building Material Dealers',
    '423510': 'Metal Service Centers',
    '423720': 'Plumbing / Heating Supplies',
    '423610': 'Electrical Equipment',
    '423490': 'Professional Equipment',
    '339999': 'Miscellaneous Manufacturing',
    '325412': 'Pharmaceutical Prep',
    '424210': 'Drugs / Druggists Sundries',
    '541330': 'Engineering Services',
}

# Priority lanes — these are DDI's strongest plays for prime/sub
PRIORITY_LANES = {
    'Facilities Support', 'Janitorial', 'Landscaping / Grounds',
    'Courier / Express Delivery', 'Local Messenger / Courier',
    'Medical Labs / Drug Testing', 'Temp Staffing', 'Business Support',
    'Office Admin', 'Freight Transport Arrangement',
    'Industrial Supplies (Wholesale)', 'Security Guards',
}


class ContractIntelligenceEngine:
    """
    Ingests, cross-references, and scores contract intelligence for DDI.
    Persists parsed data to local JSON to avoid re-parsing on every request.
    """

    def __init__(self):
        self._data = None
        self._airtable_api = None
        self._base_id = os.environ.get('AIRTABLE_BASE_ID', '')

    def _get_airtable(self):
        if self._airtable_api is None:
            from pyairtable import Api
            self._airtable_api = Api(os.environ.get('AIRTABLE_API_KEY', ''))
        return self._airtable_api

    # ------------------------------------------------------------------
    # DATA PERSISTENCE
    # ------------------------------------------------------------------

    def _load_data(self) -> Dict:
        if self._data is not None:
            return self._data
        if INTEL_DATA_PATH.exists():
            with open(INTEL_DATA_PATH, 'r') as f:
                self._data = json.load(f)
        else:
            self._data = {
                'expiring_contracts': [],
                'primes': [],
                'subs': [],
                'cross_references': [],
                'last_ingested': None,
                'sources': [],
            }
        return self._data

    def _save_data(self):
        if self._data:
            with open(INTEL_DATA_PATH, 'w') as f:
                json.dump(self._data, f, indent=2, default=str)

    # ------------------------------------------------------------------
    # INGEST — Parse raw files into structured intelligence
    # ------------------------------------------------------------------

    def ingest_all(self, folder_path: str = None) -> Dict:
        """
        Scan a folder for intelligence files and parse them all.
        Returns summary of what was ingested.
        """
        if folder_path is None:
            candidates = list(BIDS_ROOT.rglob('GOVCON_GIANTS*'))
            if candidates:
                folder_path = str(candidates[0])
            else:
                return {'error': 'No GOVCON_GIANTS folder found'}

        folder = Path(folder_path)
        if not folder.exists():
            return {'error': f'Folder not found: {folder_path}'}

        data = self._load_data()
        summary = {'files_processed': [], 'errors': []}

        for fpath in sorted(folder.iterdir()):
            fname = fpath.name.lower()
            try:
                if 'expiring' in fname and fpath.suffix == '.xlsx':
                    count = self._ingest_expiring_contracts(fpath)
                    summary['files_processed'].append({
                        'file': fpath.name, 'type': 'expiring_contracts', 'records': count
                    })
                elif 'prime' in fname and fpath.suffix == '.pdf':
                    count = self._ingest_directory_pdf(fpath, 'primes')
                    summary['files_processed'].append({
                        'file': fpath.name, 'type': 'prime_directory', 'records': count
                    })
                elif 'sub' in fname and fpath.suffix == '.pdf':
                    count = self._ingest_directory_pdf(fpath, 'subs')
                    summary['files_processed'].append({
                        'file': fpath.name, 'type': 'sub_directory', 'records': count
                    })
                elif fpath.suffix == '.csv':
                    count = self._ingest_csv(fpath)
                    summary['files_processed'].append({
                        'file': fpath.name, 'type': 'csv_data', 'records': count
                    })
            except Exception as e:
                summary['errors'].append({'file': fpath.name, 'error': str(e)})

        data['last_ingested'] = datetime.now().isoformat()
        data['sources'].append({
            'folder': str(folder),
            'ingested_at': data['last_ingested'],
            'summary': summary,
        })

        self._build_cross_references()
        self._save_data()

        summary['total_expiring'] = len(data['expiring_contracts'])
        summary['total_primes'] = len(data['primes'])
        summary['total_subs'] = len(data['subs'])
        summary['total_cross_refs'] = len(data['cross_references'])
        return summary

    def _ingest_expiring_contracts(self, xlsx_path: Path) -> int:
        """Parse expiring contracts XLSX into structured records."""
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
        ws = wb.active
        data = self._load_data()
        records = []

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row[0]:
                continue
            award_id = str(row[0] or '').strip()
            agency = str(row[1] or '').strip()
            office = str(row[2] or '').strip()
            recipient = str(row[3] or '').strip()
            naics_full = str(row[4] or '').strip()
            total_val = float(row[5]) if row[5] else 0
            start_date = str(row[6] or '')[:10]
            expiry_date = str(row[7] or '')[:10]

            naics_code = naics_full.split(' ')[0].strip() if naics_full else ''
            naics_desc = naics_full.split(' - ', 1)[1].strip() if ' - ' in naics_full else ''

            lane = self._match_ddi_lane(naics_code)

            records.append({
                'id': hashlib.md5(award_id.encode()).hexdigest()[:12],
                'award_id': award_id,
                'agency': agency,
                'office': office,
                'incumbent': recipient,
                'naics_code': naics_code,
                'naics_desc': naics_desc,
                'total_value': total_val,
                'start_date': start_date,
                'expiry_date': expiry_date,
                'ddi_lane': lane,
                'is_ddi_lane': lane is not None,
                'is_priority_lane': lane in PRIORITY_LANES if lane else False,
            })

        data['expiring_contracts'] = records
        return len(records)

    def _ingest_directory_pdf(self, pdf_path: Path, target_key: str) -> int:
        """Parse prime or sub directory PDF into structured contact records."""
        import pdfplumber

        pdf = pdfplumber.open(str(pdf_path))
        records = []
        seen = set()

        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        if not row or len(row) < 4:
                            continue
                        record = self._parse_directory_row(row, target_key)
                        if record and record['company'] not in seen:
                            seen.add(record['company'])
                            records.append(record)
            else:
                text = page.extract_text() or ''
                parsed = self._parse_directory_text(text, target_key)
                for r in parsed:
                    if r['company'] not in seen:
                        seen.add(r['company'])
                        records.append(r)

        data = self._load_data()
        data[target_key] = records
        return len(records)

    def _parse_directory_row(self, row: list, role: str) -> Optional[Dict]:
        """Parse a single table row from the directory PDF."""
        if len(row) < 4:
            return None

        company = str(row[0] or '').strip()
        if not company or company.lower() in ('company', 'company name'):
            return None

        contact_name = str(row[1] or '').strip()
        email_raw = str(row[2] or '').strip()
        phone_raw = str(row[3] or '').strip()
        naics_raw = str(row[4] or '').strip() if len(row) > 4 else ''

        email = self._extract_email(email_raw)
        phone = self._extract_phone(phone_raw)
        naics_codes = self._extract_naics_codes(naics_raw)

        return {
            'id': hashlib.md5(company.encode()).hexdigest()[:12],
            'company': company,
            'contact_name': self._clean_contact_name(contact_name),
            'email': email,
            'phone': phone,
            'naics_codes': naics_codes,
            'naics_descriptions': naics_raw,
            'role': role,  # 'primes' or 'subs'
            'ddi_lane_matches': [
                DDI_SERVICE_LANES.get(n, '') for n in naics_codes
                if n in DDI_SERVICE_LANES
            ],
            'contacted': False,
            'contact_date': None,
            'notes': '',
        }

    def _parse_directory_text(self, text: str, role: str) -> List[Dict]:
        """Fallback: parse directory from raw text when tables fail."""
        records = []
        lines = text.split('\n')

        current_company = None
        current_contact = None
        current_email = None
        current_phone = None
        current_naics = []

        for line in lines:
            line = line.strip()
            if not line or 'Contact Directory' in line or 'December 2025' in line:
                continue
            if line.startswith('Company') and ('SBLO' in line or 'Contact' in line):
                continue
            if line in ('PHONE ONLY', 'EMAIL ONLY', 'EMAIL + PHONE'):
                if current_company:
                    email = self._extract_email(current_email or '')
                    phone = self._extract_phone(current_phone or '')
                    naics_codes = []
                    for n in current_naics:
                        code = n.split(' ')[0].strip()
                        if code.isdigit():
                            naics_codes.append(code)

                    records.append({
                        'id': hashlib.md5(current_company.encode()).hexdigest()[:12],
                        'company': current_company,
                        'contact_name': self._clean_contact_name(current_contact or ''),
                        'email': email,
                        'phone': phone,
                        'naics_codes': naics_codes,
                        'naics_descriptions': '; '.join(current_naics),
                        'role': role,
                        'ddi_lane_matches': [
                            DDI_SERVICE_LANES.get(n, '') for n in naics_codes
                            if n in DDI_SERVICE_LANES
                        ],
                        'contacted': False,
                        'contact_date': None,
                        'notes': '',
                    })
                current_company = None
                current_contact = None
                current_email = None
                current_phone = None
                current_naics = []
                continue

            naics_match = re.match(r'^(\d{6})\s*-\s*(.+)', line)
            if naics_match:
                current_naics.append(line)
                continue

            if line.startswith('+ ') and 'more' in line:
                continue

            email_in_line = self._extract_email(line)
            phone_in_line = self._extract_phone(line)

            if email_in_line and not current_email:
                current_email = line
            elif phone_in_line and not current_phone:
                current_phone = line
            elif not current_company:
                parts = line.split(maxsplit=2)
                if len(parts) >= 1 and parts[0].isupper():
                    current_company = line
            elif not current_contact:
                current_contact = line

        return records

    # ------------------------------------------------------------------
    # CROSS-REFERENCE ENGINE
    # ------------------------------------------------------------------

    def _build_cross_references(self):
        """
        Match primes from the directory against expiring contract incumbents.
        This is the power move — tells DDI exactly which primes to approach.
        """
        data = self._load_data()
        cross_refs = []

        prime_names = {}
        for p in data.get('primes', []):
            name_key = self._normalize_company_name(p['company'])
            prime_names[name_key] = p

        ddi_contracts = [
            c for c in data.get('expiring_contracts', [])
            if c.get('is_ddi_lane')
        ]

        for contract in ddi_contracts:
            incumbent_key = self._normalize_company_name(contract['incumbent'])

            matched_prime = None
            for pname, pdata in prime_names.items():
                if self._fuzzy_company_match(incumbent_key, pname):
                    matched_prime = pdata
                    break

            avenues = self._score_three_avenues(contract, matched_prime)

            cross_refs.append({
                'contract': contract,
                'matched_prime': matched_prime,
                'has_sblo_contact': matched_prime is not None,
                'avenues': avenues,
                'priority_score': avenues['composite_score'],
                'recommended_action': avenues['recommended_action'],
            })

        cross_refs.sort(key=lambda x: x['priority_score'], reverse=True)
        data['cross_references'] = cross_refs

    def _get_scoring_weights(self) -> Dict:
        """Load learned weights if available, otherwise use baseline."""
        try:
            from nexus_learning_engine import get_engine
            return get_engine().get_weights('intelligence')
        except Exception:
            return {
                'sub_under_prime': {
                    'has_contact_info': 40, 'prime_in_directory': 20,
                    'is_priority_lane': 20, 'value_under_50m': 20,
                    'value_under_500m': 10, 'value_over_500m': 5,
                },
                'prime_recompete': {
                    'is_priority_lane': 30, 'value_under_10m': 30,
                    'value_under_50m': 20, 'value_under_200m': 10,
                    'value_over_200m': 5, 'va_agency': 15, 'hhs_agency': 10,
                },
                'hire_subs': {
                    'priority_lane': 25, 'value_under_25m': 25,
                    'value_under_100m': 15, 'high_sub_availability': 20,
                },
            }

    def _score_three_avenues(self, contract: Dict, prime: Optional[Dict]) -> Dict:
        """
        Score a contract for DDI's three avenues using LEARNED weights.
        Weights start as baseline and evolve as Dee logs actions and outcomes.
        """
        w = self._get_scoring_weights()
        w_sub = w.get('sub_under_prime', {})
        w_prime = w.get('prime_recompete', {})
        w_hire = w.get('hire_subs', {})

        scores = {
            'sub_under_prime': 0,
            'prime_recompete': 0,
            'hire_subs': 0,
        }

        val = contract.get('total_value', 0)
        is_priority = contract.get('is_priority_lane', False)
        lane = contract.get('ddi_lane', '')
        agency = contract.get('agency', '')

        # AVENUE 1: Sub under the prime (weights from learning engine)
        if prime and (prime.get('email') or prime.get('phone')):
            scores['sub_under_prime'] += w_sub.get('has_contact_info', 40)
        if prime:
            scores['sub_under_prime'] += w_sub.get('prime_in_directory', 20)
        if is_priority:
            scores['sub_under_prime'] += w_sub.get('is_priority_lane', 20)
        if val > 0:
            if val <= 50_000_000:
                scores['sub_under_prime'] += w_sub.get('value_under_50m', 20)
            elif val <= 500_000_000:
                scores['sub_under_prime'] += w_sub.get('value_under_500m', 10)
            else:
                scores['sub_under_prime'] += w_sub.get('value_over_500m', 5)

        # AVENUE 2: Prime the recompete (weights from learning engine)
        if is_priority:
            scores['prime_recompete'] += w_prime.get('is_priority_lane', 30)
        if val <= 10_000_000:
            scores['prime_recompete'] += w_prime.get('value_under_10m', 30)
        elif val <= 50_000_000:
            scores['prime_recompete'] += w_prime.get('value_under_50m', 20)
        elif val <= 200_000_000:
            scores['prime_recompete'] += w_prime.get('value_under_200m', 10)
        else:
            scores['prime_recompete'] += w_prime.get('value_over_200m', 5)
        if 'Veterans' in agency:
            scores['prime_recompete'] += w_prime.get('va_agency', 15)
        elif 'Health' in agency:
            scores['prime_recompete'] += w_prime.get('hhs_agency', 10)

        # AVENUE 3: Hire subs (weights from learning engine)
        if lane in PRIORITY_LANES:
            scores['hire_subs'] += w_hire.get('priority_lane', 25)
        if val <= 25_000_000:
            scores['hire_subs'] += w_hire.get('value_under_25m', 25)
        elif val <= 100_000_000:
            scores['hire_subs'] += w_hire.get('value_under_100m', 15)
        if any(k in (lane or '') for k in ('Janitorial', 'Grounds', 'Facilities')):
            scores['hire_subs'] += w_hire.get('high_sub_availability', 20)

        composite = max(scores.values())
        best_avenue = max(scores, key=scores.get)

        action_map = {
            'sub_under_prime': 'Contact SBLO — get on their sub list',
            'prime_recompete': 'Monitor SAM.gov for recompete — prepare to bid as prime',
            'hire_subs': 'Build sub team now — bid as prime when opportunity drops',
        }

        return {
            **scores,
            'composite_score': composite,
            'best_avenue': best_avenue,
            'recommended_action': action_map.get(best_avenue, ''),
        }

    # ------------------------------------------------------------------
    # QUERY — What the dashboard and API call
    # ------------------------------------------------------------------

    def get_pipeline(self, lane_filter: str = None, min_score: int = 0) -> Dict:
        """
        Get the full three-avenue pipeline for the dashboard.
        Returns categorized opportunities ready for display.
        """
        data = self._load_data()
        cross_refs = data.get('cross_references', [])

        if not cross_refs:
            self.ingest_all()
            data = self._load_data()
            cross_refs = data.get('cross_references', [])

        if lane_filter:
            cross_refs = [
                cr for cr in cross_refs
                if lane_filter.lower() in (cr['contract'].get('ddi_lane', '') or '').lower()
            ]

        if min_score > 0:
            cross_refs = [cr for cr in cross_refs if cr['priority_score'] >= min_score]

        sub_now = []
        prime_later = []
        hire_subs = []

        for cr in cross_refs:
            avenue = cr['avenues']['best_avenue']
            item = self._format_pipeline_item(cr)
            if avenue == 'sub_under_prime':
                sub_now.append(item)
            elif avenue == 'prime_recompete':
                prime_later.append(item)
            else:
                hire_subs.append(item)

        by_lane = defaultdict(int)
        by_agency = defaultdict(int)
        for cr in cross_refs:
            lane = cr['contract'].get('ddi_lane', 'Other')
            agency = cr['contract'].get('agency', 'Unknown')
            by_lane[lane] += 1
            by_agency[agency] += 1

        return {
            'last_ingested': data.get('last_ingested'),
            'total_opportunities': len(cross_refs),
            'avenues': {
                'sub_under_prime': {
                    'title': 'Sub Under Prime NOW',
                    'subtitle': 'Get on their vendor list — immediate revenue path',
                    'color': 'blue',
                    'count': len(sub_now),
                    'items': sub_now[:50],
                },
                'prime_recompete': {
                    'title': 'Prime the Recompete',
                    'subtitle': 'Contract expiring — DDI bids as prime next time',
                    'color': 'purple',
                    'count': len(prime_later),
                    'items': prime_later[:50],
                },
                'hire_subs': {
                    'title': 'Build Sub Team & Prime',
                    'subtitle': 'DDI primes, hires Tier 2 subs to perform',
                    'color': 'amber',
                    'count': len(hire_subs),
                    'items': hire_subs[:50],
                },
            },
            'stats': {
                'total_expiring': len(data.get('expiring_contracts', [])),
                'ddi_lane_matches': len([
                    c for c in data.get('expiring_contracts', []) if c.get('is_ddi_lane')
                ]),
                'primes_in_directory': len(data.get('primes', [])),
                'subs_in_directory': len(data.get('subs', [])),
                'cross_ref_matches': len([
                    cr for cr in cross_refs if cr.get('has_sblo_contact')
                ]),
                'by_lane': dict(sorted(by_lane.items(), key=lambda x: x[1], reverse=True)),
                'by_agency': dict(sorted(by_agency.items(), key=lambda x: x[1], reverse=True)[:10]),
            },
        }

    def get_expiring_contracts(self, ddi_only: bool = True, lane: str = None) -> List[Dict]:
        """Get expiring contracts, optionally filtered to DDI lanes."""
        data = self._load_data()
        contracts = data.get('expiring_contracts', [])
        if ddi_only:
            contracts = [c for c in contracts if c.get('is_ddi_lane')]
        if lane:
            contracts = [c for c in contracts if lane.lower() in (c.get('ddi_lane', '') or '').lower()]
        contracts.sort(key=lambda c: c.get('total_value', 0), reverse=True)
        return contracts

    def get_prime_directory(self, lane_filter: str = None) -> List[Dict]:
        """Get primes from directory, optionally filtered by DDI lane match."""
        data = self._load_data()
        primes = data.get('primes', [])
        if lane_filter:
            primes = [
                p for p in primes
                if any(lane_filter.lower() in m.lower() for m in p.get('ddi_lane_matches', []))
            ]
        return primes

    def get_sub_directory(self, lane_filter: str = None) -> List[Dict]:
        """Get Tier 2 subs, optionally filtered by lane."""
        data = self._load_data()
        subs = data.get('subs', [])
        if lane_filter:
            subs = [
                s for s in subs
                if any(lane_filter.lower() in m.lower() for m in s.get('ddi_lane_matches', []))
            ]
        return subs

    def get_priority_outreach(self, limit: int = 20) -> List[Dict]:
        """
        Top priority outreach targets — primes with expiring contracts
        in DDI's lanes who have SBLO contact info.
        """
        data = self._load_data()
        cross_refs = data.get('cross_references', [])

        priority = [
            cr for cr in cross_refs
            if cr.get('has_sblo_contact') and cr.get('priority_score', 0) >= 50
        ]
        priority.sort(key=lambda x: x['priority_score'], reverse=True)

        return [self._format_pipeline_item(cr) for cr in priority[:limit]]

    # ------------------------------------------------------------------
    # TASK GENERATION — Create actionable tasks in Airtable
    # ------------------------------------------------------------------

    def generate_outreach_tasks(self, avenue: str = 'sub_under_prime', limit: int = 10) -> Dict:
        """
        Generate Airtable TASKS for priority outreach based on intelligence.
        Only creates tasks for contacts not yet reached out to.
        """
        data = self._load_data()
        cross_refs = data.get('cross_references', [])

        targets = [
            cr for cr in cross_refs
            if cr['avenues']['best_avenue'] == avenue
            and cr.get('has_sblo_contact')
            and cr.get('priority_score', 0) >= 40
        ]
        targets.sort(key=lambda x: x['priority_score'], reverse=True)
        targets = targets[:limit]

        created = []
        api = self._get_airtable()

        for cr in targets:
            prime = cr.get('matched_prime', {})
            contract = cr['contract']

            if avenue == 'sub_under_prime':
                title = f"SBLO Outreach: {prime.get('company', 'Unknown')} — {contract.get('ddi_lane', '')}"
                desc_parts = []
                if prime.get('contact_name'):
                    desc_parts.append(f"Contact: {prime['contact_name']}")
                if prime.get('email'):
                    desc_parts.append(f"TO: {prime['email']}")
                if prime.get('phone'):
                    desc_parts.append(f"Phone: {prime['phone']}")
                desc_parts.append(f"Contract: {contract.get('award_id', '')} ({contract.get('agency', '')})")
                desc_parts.append(f"Value: ${contract.get('total_value', 0)/1e6:,.1f}M")
                desc_parts.append(f"Expires: {contract.get('expiry_date', '')}")
                desc_parts.append(f"Lane: {contract.get('ddi_lane', '')}")
                desc_parts.append(f"Avenue: Sub under prime — get on vendor list")
                description = ' | '.join(desc_parts)
                priority = 'HIGH'
                project = 'SBLO Outreach'

            elif avenue == 'prime_recompete':
                title = f"Monitor Recompete: {contract.get('award_id', '')} — {contract.get('ddi_lane', '')}"
                desc_parts = [
                    f"Agency: {contract.get('agency', '')}",
                    f"Incumbent: {contract.get('incumbent', '')}",
                    f"Value: ${contract.get('total_value', 0)/1e6:,.1f}M",
                    f"Expires: {contract.get('expiry_date', '')}",
                    f"Lane: {contract.get('ddi_lane', '')}",
                    f"Avenue: Watch SAM.gov for recompete — prepare to bid as prime",
                ]
                description = ' | '.join(desc_parts)
                priority = 'MEDIUM'
                project = 'Recompete Pipeline'

            else:
                title = f"Build Sub Team: {contract.get('ddi_lane', '')} — {contract.get('agency', '')}"
                desc_parts = [
                    f"Contract: {contract.get('award_id', '')}",
                    f"Agency: {contract.get('agency', '')}",
                    f"Value: ${contract.get('total_value', 0)/1e6:,.1f}M",
                    f"Lane: {contract.get('ddi_lane', '')}",
                    f"Avenue: DDI primes, hire Tier 2 subs from directory",
                ]
                description = ' | '.join(desc_parts)
                priority = 'MEDIUM'
                project = 'Sub Team Building'

            try:
                record = api.table(self._base_id, 'TASKS').create({
                    'TITLE': title,
                    'DESCRIPTION': description,
                    'PRIORITY': priority,
                    'STATUS': 'TO DO',
                    'PROJECTS': project,
                })
                created.append({
                    'record_id': record['id'],
                    'title': title,
                    'priority': priority,
                })
            except Exception as e:
                created.append({'title': title, 'error': str(e)})

        return {
            'avenue': avenue,
            'tasks_created': len([c for c in created if 'record_id' in c]),
            'tasks': created,
        }

    # ------------------------------------------------------------------
    # HELPERS
    # ------------------------------------------------------------------

    def _format_pipeline_item(self, cross_ref: Dict) -> Dict:
        """Format a cross-reference into a dashboard-ready item."""
        contract = cross_ref['contract']
        prime = cross_ref.get('matched_prime') or {}
        avenues = cross_ref['avenues']

        return {
            'id': contract.get('id', ''),
            'award_id': contract.get('award_id', ''),
            'agency': contract.get('agency', ''),
            'incumbent': contract.get('incumbent', ''),
            'lane': contract.get('ddi_lane', ''),
            'value': contract.get('total_value', 0),
            'value_display': f"${contract.get('total_value', 0)/1e6:,.1f}M" if contract.get('total_value', 0) > 0 else 'N/A',
            'expiry': contract.get('expiry_date', ''),
            'is_priority_lane': contract.get('is_priority_lane', False),
            'prime_company': prime.get('company', ''),
            'prime_contact': prime.get('contact_name', ''),
            'prime_email': prime.get('email', ''),
            'prime_phone': prime.get('phone', ''),
            'has_contact': bool(prime.get('email') or prime.get('phone')),
            'best_avenue': avenues.get('best_avenue', ''),
            'score': avenues.get('composite_score', 0),
            'sub_score': avenues.get('sub_under_prime', 0),
            'prime_score': avenues.get('prime_recompete', 0),
            'hire_score': avenues.get('hire_subs', 0),
            'action': avenues.get('recommended_action', ''),
        }

    def _match_ddi_lane(self, naics_code: str) -> Optional[str]:
        """Check if a NAICS code falls in DDI's service lanes."""
        if naics_code in DDI_SERVICE_LANES:
            return DDI_SERVICE_LANES[naics_code]
        for code, lane in DDI_SERVICE_LANES.items():
            if naics_code.startswith(code[:4]):
                return lane
        return None

    def _normalize_company_name(self, name: str) -> str:
        """Normalize company name for matching."""
        name = name.upper().strip()
        for suffix in [', INC.', ', INC', ' INC.', ' INC', ', LLC', ' LLC',
                       ' CORPORATION', ' CORP.', ' CORP', ' CO.', ' CO',
                       ', L.L.C.', ' L.L.C.', ' LTD', ', LTD',
                       ' SERVICES', ' SOLUTIONS', ' GROUP']:
            name = name.replace(suffix, '')
        return name.strip()

    def _fuzzy_company_match(self, name1: str, name2: str) -> bool:
        """Fuzzy match two normalized company names."""
        if name1 == name2:
            return True
        if name1 in name2 or name2 in name1:
            return True
        words1 = set(name1.split())
        words2 = set(name2.split())
        if len(words1) > 1 and len(words2) > 1:
            overlap = words1 & words2
            if len(overlap) >= min(len(words1), len(words2)) * 0.7:
                return True
        return False

    def _extract_email(self, text: str) -> str:
        """Extract first valid email from messy PDF text."""
        match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
        return match.group(0).lower() if match else ''

    def _extract_phone(self, text: str) -> str:
        """Extract first phone number from text."""
        match = re.search(r'[\(]?\d{3}[\)]?[\s.-]?\d{3}[\s.-]?\d{4}', text)
        return match.group(0) if match else ''

    def _extract_naics_codes(self, text: str) -> List[str]:
        """Extract NAICS codes from text."""
        codes = re.findall(r'\b(\d{6})\b', text)
        return list(set(codes))

    def _clean_contact_name(self, name: str) -> str:
        """Clean up a contact name from messy PDF parsing."""
        name = re.sub(r'[;,].*', '', name)
        name = re.sub(r'\b(SBLO|Website|Small|Email)\b', '', name, flags=re.IGNORECASE)
        name = re.sub(r'\s+', ' ', name).strip()
        return name

    def _ingest_csv(self, csv_path: Path) -> int:
        """Ingest a CSV file into the data store."""
        data = self._load_data()
        records = []
        with open(csv_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f)
            for row in reader:
                records.append(dict(row))
        key = csv_path.stem.replace('-', '_').replace(' ', '_').lower()
        data[key] = records
        return len(records)


# ------------------------------------------------------------------
# MODULE-LEVEL CONVENIENCE FUNCTIONS (called by api_server.py)
# ------------------------------------------------------------------

_engine = None

def get_engine() -> ContractIntelligenceEngine:
    global _engine
    if _engine is None:
        _engine = ContractIntelligenceEngine()
    return _engine

def handle_ingest(folder_path: str = None) -> Dict:
    return get_engine().ingest_all(folder_path)

def handle_get_pipeline(lane: str = None, min_score: int = 0) -> Dict:
    return get_engine().get_pipeline(lane, min_score)

def handle_get_expiring(ddi_only: bool = True, lane: str = None) -> List[Dict]:
    return get_engine().get_expiring_contracts(ddi_only, lane)

def handle_get_primes(lane: str = None) -> List[Dict]:
    return get_engine().get_prime_directory(lane)

def handle_get_subs(lane: str = None) -> List[Dict]:
    return get_engine().get_sub_directory(lane)

def handle_get_priority_outreach(limit: int = 20) -> List[Dict]:
    return get_engine().get_priority_outreach(limit)

def handle_generate_tasks(avenue: str = 'sub_under_prime', limit: int = 10) -> Dict:
    return get_engine().generate_outreach_tasks(avenue, limit)
